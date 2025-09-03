import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
import copy

class ScheduleExcel:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self._load()

    def _load(self):
        self.book = load_workbook(self.xlsx_path)
        # No in-memory caching of sheets as dataframes; read on demand

    def list_doctors(self):
        return self.book.sheetnames

    def upcoming_days(self, n=7):
        base = date.today()
        return [(base + pd.Timedelta(days=i)).isoformat() for i in range(n)]

    def available_slots(self, doctor, target_date_iso):
        df = pd.read_excel(self.xlsx_path, sheet_name=doctor, parse_dates=["date"])
        target = pd.to_datetime(target_date_iso).date()
        rows = df[df['date'].dt.date == target]
        avail = rows[rows['status'].str.lower()=="available"]
        out = []
        for _, r in avail.iterrows():
            out.append({
                "date": r['date'].date(),
                "start_time": r['start_time'],
                "end_time": r['end_time'],
                "slot_length": int(r['slot_length'])
            })
        return out

    def find_slots(self, doctor, required_minutes):
        """Return list of slot dicts available (first-fit)"""
        df = pd.read_excel(self.xlsx_path, sheet_name=doctor, parse_dates=["date"])
        # iterate across dates ascending
        df = df.sort_values(["date","start_time"])
        slots = []
        for _, r in df.iterrows():
            if str(r['status']).lower() != "available":
                continue
            if int(r['slot_length']) >= required_minutes:
                slots.append({
                    "date": r['date'].date(),
                    "start_time": r['start_time'],
                    "end_time": r['end_time'],
                    "slot_length": int(r['slot_length'])
                })
        return slots

    def book_slot(self, doctor, slot):
        """Mark the first matching slot as Booked and save workbook. Return True/False."""
        df = pd.read_excel(self.xlsx_path, sheet_name=doctor, parse_dates=["date"])
        mask = (
            (df['date'].dt.date == slot['date']) &
            (df['start_time'] == slot['start_time']) &
            (df['end_time'] == slot['end_time'])
        )
        if mask.sum() == 0:
            return False
        if df.loc[mask, 'status'].iloc[0].lower() != "available":
            return False
        df.loc[mask, 'status'] = "Booked"
        # write back to excel safely
        with pd.ExcelWriter(self.xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=doctor, index=False)
        return True
